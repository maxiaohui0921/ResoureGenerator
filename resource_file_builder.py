#-*-coding:utf-8-*-
#__author__='maxiaohui'

'''
As in resource files, text use variable which value is saved in Data Source(Excel)
The script is used to build available resource file
'''

import codecs
import openpyxl,time,os,shutil,re,sys
from openpyxl.styles import Font,colors
import smtplib
from email.mime.multipart import MIMEMultipart
from email.header import Header
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

reload(sys)
sys.setdefaultencoding( "utf-8" )

#配置文件  1、产品提供的数据源 2、开发提供的template 3、结果反馈文件
# localPath=os.path.abspath('.')   #当前文件的上一级目录
localPath=os.path.dirname(__file__)

#源文件目录
excel_source= os.path.join(localPath,"Bbox_fontDB.xlsx")
language_column={"cn":3,"en":4,"japan":5}
#反馈
# excel_elementnotfound=os.path.join(localPath,'notFount_%s.xlsx'%time.strftime('%Y%m%d%H%M%S',time.localtime(time.time())))
excel_config=os.path.join(localPath,'config.xlsx')
#android项目目录
projectPath=os.path.join(localPath,"packages","packages")

#邮件信息配置
sender='maxiaohui@beeboxes.com'
to_receiver=['maxiaohui@beeboxes.com']
cc_reciver=['maxiaohui0921@163.com']
receiver = to_receiver + cc_reciver
email_text = '''
        Hi all
        ----------本邮件由Python2.7脚本自动发送----------
        附件中是资源文件build时缺少的字段，请尽快补齐。
        
        备注：以下是自动生成的资源文件：  
        '''

class easyExcel():
    '''
    excel基本操作，读，写，搜索
    '''

    # 打开文件或者新建文件（如果不存在的话）
    def __init__(self, filename=None):
        self.filename = filename
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            wb.save(filename=filename)
        self.xlBook = openpyxl.load_workbook(self.filename,data_only=True)

    # 加入一个指定名称的sheet，返回sheet对象
    def addSheet(self, sheetname):
        self.sheetname = sheetname
        self.xlBook.create_sheet(self.sheetname)
        xlSheet = self.xlBook.get_sheet_by_name(self.sheetname)
        return xlSheet

    # 根据名称获得某个sheet，返回sheet对象
    def getSheet(self, sheetname):
        self.sheetname = sheetname
        xlSheet = self.xlBook.get_sheet_by_name(self.sheetname)
        return xlSheet

    #删除sheet
    def delSheet(self,sheetname):
        self.xlBook.remove(self.getSheet(sheetname))

    # 写cell
    def writeCell(self, xlSheet, r, c, v):
        self.xlSheet = xlSheet
        # time.sleep(0.02)
        self.xlSheet.cell(r, c).value = v

    # 获取cell
    def getCell(self, xlSheet, r, c):
        return xlSheet.cell(r, c).value

    # 把列表写入指定行,可指定起始列
    def writeRow(self, xlSheet, list, rowID, startColumn=1):
        for i in range(startColumn, startColumn + len(list)):
            self.writeCell(xlSheet, rowID, i, list[i - startColumn])

    # 把列表制定行的值获取回来，存成列表
    def getRow_value(self, xlSheet, rowID, columnEnd, columnStart=1):
        rowList=[]
        for column in range(columnStart,columnEnd+1):
            rowList.append(self.getCell(xlSheet,rowID,column))
        return rowList

    # 筛选出某一列包含某字符串的条目的个数
    def filterColumnTextContains(self, xlSheet, rowNumbers, columnNumber, filterChar):
        count = 0
        rowList = []
        for i in range(2, rowNumbers + 1):
            if xlSheet.cell(i, columnNumber).value.find(filterChar) >= 0:
                count += 1
                rowList.append(i)
        return count, rowList  # 返回筛选出的数目，以及行列表

    # 在某一列搜索某个字符串，返回行数
    def searchTextByColumn(self, sheet, column, endRow, txt,startRow=2):
        rowNums = []
        for i in range(startRow, endRow + 1):
            cellValue = str(sheet.cell(i, column).value).strip()  #获得字符串内容后，先字符串化，移除前后空格
            if cellValue!=None:   #有时候是空， None
                cellValue=cellValue.replace("&#12288;"," ")  #从excel获取的字符串有java空白符
            if txt!=None:
                txt=txt.replace(u"\u3000",u" ")   #从xml获取到的字符串，里面含有全角的空白符
            if cellValue == txt:
                rowNums.append(i)
        return rowNums

    #查找列表的合集，返回列表
    def getSameText(self,list1, list2):
        sameList = []
        if len(list1)>=1 and len(list2)>=1:
            for i in list1:
                if i in list2:
                    sameList.append(i)
        if len(sameList)==0:  #无共同值时，以list2为准，返回list2
            sameList=list2
        return sameList

    # 联合搜索
    def searchTextByMultiColumn(self,sheet,endRow,startRow=2,*args):   # ([text,column],[text,column])
        rowLists=[]
        for i in args:
            rowNums=self.searchTextByColumn(sheet,i[1],endRow,i[0],startRow)
            rowLists.append(rowNums)
        print(rowLists)
        for index in range(len(rowLists)-1):
            sameList=self.getSameText(rowLists[index],rowLists[index+1])
        print(sameList)
        if not sameList:
            sameList.append(0)
            print(sameList)
        return sameList[0]   #返回行数

    #获取行数，列数
    def getRows_numbers(self, xl_sheet):
        row = xl_sheet.max_row
        column = xl_sheet.max_column
        return row,column

    #设置字体颜色，默认值是黑色
    def setFontColor(self,r,c,color=colors.BLACK):
        a1 = self.xlSheet.cell(r, c)
        ft = Font(color=color)
        a1.font = ft

    #设置列的宽度
    def setColumnWidth(self,cList,widthList):
        for cindex in range(len(cList)):
            self.xlSheet.column_dimensions[cList[cindex]].width = widthList[cindex]

    # 保存当前文件
    def save(self):
        self.xlBook.save(self.filename)

    # 关闭excel应用
    def closeFile(self):
        self.xlBook.close()

class fileHandler():
    '''
    文件的copy，复制，删除，移动等
    '''
    def copy_file(self,source,aimpath,postfix=""):  #把源文件copy到目标目录下，文件名不变或加后缀
        path,filename=os.path.split(source)
        try:
            filename=filename.split(".")[0]+postfix+'.'+filename.split(".")[1]
        except IndexError:   #处理没有后缀名的配置文件
            pass
        copyed_file = os.path.join(aimpath, filename)
        shutil.copy(source,os.path.join(aimpath,filename))
        return copyed_file

    def move_file(self,source,aim):
        shutil.move(source,aim)

    def del_file(self,source):
        os.remove(source)

    def del_folder(self,folderPath): #删除空文件夹
        os.rmdir(folderPath)

    def make_dir(self,path):  #创建一个文件夹
        os.mkdir(path)

class resourceBuilder():  #一个类对应一个template file，多种语言
    '''
    支持替换文件中的变量T000123，生成目标resource文件,自动反馈问题结果
    '''
    fHandler=fileHandler()

    #以下是初始化部分，准备各种文件
    def __init__(self,configExcel,sourceExcel):  #传入参数：配置文件xml，源文件（xuliyang提供)
        self.config=configExcel
        self.source=sourceExcel
        self.result_folder = self.prepare_result_folder()
        self.prepare_source_feedback_excel()
        self.config_list = self.get_config()

    def prepare_source_feedback_excel(self): #准备配置文件，源文件，feedback文件
        self.ex_config = easyExcel(self.config)   #配置文件，只用来读取。 增加语言或者增加模板直接新加一行
        self.config_sht = self.ex_config.getSheet("Template")
        self.row_config, self.column_config = self.ex_config.getRows_numbers(self.config_sht)

        self.ex = easyExcel(self.source)            #源文件，只用来读取
        self.source_sht = self.ex.getSheet("Sheet1")
        self.row, self.column = self.ex.getRows_numbers(self.source_sht)

        self.excel_feedback = os.path.join(self.result_folder, 'Resource_notFound_%s.xlsx' % time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
        self.ex_feedback = easyExcel(self.excel_feedback)  #反馈给xuliyang的文件，只用来读写
        self.ex_feedback.delSheet("Sheet")
        self.feedback_sheet = self.ex_feedback.addSheet("result")
        self.ex_feedback.writeRow(self.feedback_sheet,['文件模板名','行数','问题行','备注','语言'],1)
        self.ex_feedback.setColumnWidth(["A","B","C","D","E"],[80,5,65,50,5])
        self.row_feedback=2

    def get_config(self):   #从配置文件中获取配置列表
        config_list=[]
        row_number,column_number=self.ex_config.getRows_numbers(self.config_sht)
        for row in range(2,row_number+1):
            rowValueList = self.ex_config.getRow_value(self.config_sht,row,column_number)
            d=dict(zip(['template_path','language','aim_path','app_name'],rowValueList))
            config_list.append(d)
        return config_list

    def prepare_result_folder(self):  #生成结果文件夹
        # os.chdir(projectPath)
        result_folder = os.path.join(localPath, "temparory_folder")
        result_folder = os.path.join(localPath, "result")
        try:
            os.mkdir(result_folder)
        except OSError:
            shutil.rmtree(result_folder)
            os.mkdir(result_folder)
        return result_folder

    #以下部分为主流程，获取待处理文件，根据变量生成目标文件，放到制定目录中
    def get_files(self,template_file,result_folder): #template_path, aim_path来自于配置文件  路径完整
        template_file=projectPath+template_file
        template_file = self.fHandler.copy_file(template_file,result_folder)
        return template_file    #处理过程中的文件

    #暂时不用，一旦模板和目标文件错行，不好找，工作量大
    def compare(self,line,line_template):  #对比两个字符串，找到中间不同的地方，返回value值
        forward,backward = 0,-1
        while True:
            if line[forward]==line_template[forward]:
                forward+=1
            else:
                break
        while True:
            if line[backward]==line_template[backward]:
                backward-=1
            else:
                break
        return line_template[forward:backward]

    # 暂时不用，一旦模板和目标文件错行，不好找，工作量大
    def getValue_from_aim(self,aim_file,line_number,line_template):  #在目标文件中找到
        with open(aim_file, "r", encoding="utf-8") as f:
            line_no=0
            for line in f:
                line_no+=1
                if line_no==line_number:
                    diff_value=self.compare(line,line_template)
                    break
        return diff_value

    def getValue_from_excel(self,id,language):  # 根据id：T00054，获取value值，如果找不到，返回id
        aimrow = self.ex.searchTextByColumn(self.source_sht, 1, self.row, id)  # 第一列配置的是id
        # 处理找不到ID的情况 （理论上来说，只要写了id都能找到，但是如果不小心被误删除了）
        if aimrow:
            menuText = self.ex.getCell(self.source_sht, aimrow[0], language_column[language])
            if menuText==" ":   #处理有公式，但是没有值的数据，这时候，依然返回id
                menuText = None
        else:
            menuText = id  # 如果找不到的话，直接返回原id
        return str(menuText)

    def write_line_feedback(self,feedbackList,color):  #对于有问题的条目，写入内容，设置comments列的颜色
        self.ex_feedback.writeRow(self.feedback_sheet, feedbackList,self.row_feedback)
        self.ex_feedback.setFontColor(self.row_feedback, 4, color)
        self.row_feedback += 1

    def get_values(self,template_file,language):   #处理两种异常：有变量号找不到值，新增变量{{测试文本}}
        file_data = ""
        with codecs.open(template_file,"r","utf-8") as f:
            line_no=1
            for line in f:
                txtVa = re.findall(r"(S\d{5,10}$)", line)   #注意！变量标号的数字位数最短不能少于5位数字
                txtNewCreated = re.findall(r"{{(.*?)}}", line)
                if txtVa:   #在文件中找到变量 T00234
                    old_str = txtVa[0]
                    new_str = self.getValue_from_excel(old_str, language)
                    if old_str == new_str:  #代表excel中没有找到相关变量，记录到反馈文件中
                        feedbackList=[template_file,line_no,line, "未找到条目:%s"%new_str,language]
                        self.write_line_feedback(feedbackList,colors.RED)
                    if new_str=="None":   #代表excel中有该变量名，但是字符串为空，例如英文没有翻译
                        feedbackList=[template_file, line_no, line, "%s未翻译:%s" %(language,old_str), language]
                        self.write_line_feedback(feedbackList, colors.DARKYELLOW)
                        new_str=old_str
                    line = line.replace(old_str, new_str)
                if txtNewCreated: #是新添加的变量，中文替换，其他语言置空
                    line = line.replace("{{", "")
                    line = line.replace("}}", "")
                    feedbackList=[template_file, line_no, line, "新增条目:%s" % txtNewCreated[0], language]
                    self.write_line_feedback(feedbackList, colors.BLUE)
                    # if language!="cn":
                    #     line = line.replace(txtNewCreated[0],"")   对于新增语言不做任何处理，只是去掉大括号
                line_no += 1
                file_data += line

        with codecs.open(template_file, "w", "utf-8") as f:
            f.write(file_data)

    def clear_down(self,template_file,aim_file): #move template文件到指定目录，保存feedback文件
        self.fHandler.move_file(template_file,projectPath+aim_file)
        global email_text
        email_text=email_text+"\n"+aim_file
        print("generate resource file: %s"%(projectPath+aim_file))

    def send_email(self,sender, receiver,text,attachment):
        # 第三方 SMTP 服务
        mail_host = "fkdsjflds"  # 设置服务器
        mail_user = "mfdsf"  # 用户名
        mail_pass = "3423"  # 口令

        content = MIMEText(text,"plain","utf-8")
        message = MIMEMultipart()
        message.attach(content)
        message['From'] = sender
        message['To'] = ";".join(to_receiver)
        message['Cc'] = ";".join(cc_reciver)
        subject = '资源文件反馈_%s (临时调试，请忽略当前邮件)' % time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
        message['Subject'] = Header(subject, 'utf-8')
        xlsx = MIMEApplication(open(attachment, 'rb').read())
        xlsx["Content-Type"] = 'application/octet-stream'
        xlsx.add_header('Content-Disposition', 'attachment', filename=os.path.split(attachment)[1])
        message.attach(xlsx)
        try:
            smtpObj = smtplib.SMTP()
            smtpObj.connect(mail_host, 25)  # 25 为 SMTP 端口号
            smtpObj.login(mail_user, mail_pass)
            smtpObj.sendmail(sender, receiver, message.as_string())
            print "Send email successfuly"
        except  Exception as e:
            print(e)

    def tear_down(self,sendEmail=True):  #保存feedback文件，发送邮件，删除feedback文件，清空临时文件夹
        self.ex_feedback.save()
        if sendEmail:
            self.send_email(sender, receiver, email_text,self.excel_feedback)
        self.fHandler.del_file(self.excel_feedback)
        self.fHandler.del_folder(self.result_folder)

    def build_all_config(self): #把excel中所有的配置条目，从第二行开始，都build了
        for i in self.config_list:
            try:
                template_file = self.get_files(i['template_path'], rd.result_folder)
                self.get_values(template_file, i["language"])
                self.clear_down(template_file, i['aim_path'])
            except IOError:
                print("无模板文件：%s，不处理" % i['template_path'])
        self.tear_down()

    def build_config_column(self,column_no): #把excel某个条目，从第二行开始，都build了
        column=1
        for i in self.config_list:
            column+=1
            if column==column_no:
                try:
                    template_file = self.get_files(i['template_path'], rd.result_folder)
                    self.get_values(template_file, i["language"])
                    self.clear_down(template_file, i['aim_path'])
                except IOError:
                    print("无模板文件：%s，不处理" % i['template_path'])
        self.tear_down(sendEmail=False)

    def build_config_app(self,app_name): #针对app进行build语言文件
        for i in self.config_list:
            if i["app_name"]==app_name:
                try:
                    template_file = self.get_files(i['template_path'], rd.result_folder)
                    self.get_values(template_file, i["language"])
                    self.clear_down(template_file, i['aim_path'])
                except IOError:
                    print("无模板文件：%s，不处理" % i['template_path'])
        self.tear_down(sendEmail=False)

if __name__ == "__main__":
    rd=resourceBuilder(excel_config,excel_source)   #excel_config由开发配置， excel_source由xuliyang提供
    try:
        if sys.argv[1]:
            rd.build_config_app(sys.argv[1])       #带了app参数的话，会只build当前app
    except IndexError:
        rd.build_all_config()
    # rd.build_config_column(5)                   #如果只想生成一个文件，可以用这条语句
